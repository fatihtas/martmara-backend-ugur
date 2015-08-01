import json
import pymongo
import sys
import datetime, time
import bson
import os
import openpyxl
from bson.objectid import ObjectId
from bson import Binary, Code
from bson import json_util
from difflib import SequenceMatcher
#MmPasa_189m

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def get_venueobjarray():
	wb = openpyxl.reader.excel.load_workbook('Viveka Transferred Data.xlsx',data_only = True,guess_types=True)
	#print (wb.get_sheet_names())
	sheet_ranges = wb['Data Entry Sheet']
	allvenues = []
	for x in range(24,2122):#for each venue
		rowtuple = {}
		name = sheet_ranges['H'+str(x)].value#[,[]]
		while name.endswith(' '):
			name = name[:-1]
		rowtuple['name'] = name
		rowtuple['workinghours'] = []
		for day in range(7):#for each day
			opening = sheet_ranges[openpyxl.utils.get_column_letter(day*2+16)+str(x)].value
			closing = sheet_ranges[openpyxl.utils.get_column_letter(day*2+17)+str(x)].value
			
			if (not isinstance(opening,datetime.datetime) and not isinstance(opening,datetime.time)):
				print("bugun calismiyoruz.",opening,closing)
				continue
			if isinstance(opening,datetime.datetime):
				opening = opening.time()
			if isinstance(closing,datetime.datetime):
				closing = closing.time()	
			opening = opening.hour*60+opening.minute
			closing = closing.hour*60+closing.minute
			
			rowtuple['workinghours'].append({'day':day+1,'opening':opening,'closing':closing})
		allvenues.append(rowtuple)
	return allvenues
	
def update(db_venues,day_json,_id,dbname,excelvenuename):
	print ("updating: ",_id,'\t',dbname,'\t',excelvenuename)
	print ("with json:",day_json)
	db_venues.update({
		'_id' : _id
		},{
		'$set':{
		'operatingHours':day_json
		}
		}
		,multi=False, upsert=False
		)
		
	return
		
def get_day_json(correctvenue):
	operatingHours = []
	for eachday in correctvenue['workinghours']:
		day_json={}
		day_json['dayOfTheWeek'] = eachday['day']
		day_json['openingHour'] = eachday['opening']
		day_json['closingHour'] = eachday['closing']
		day_json['openingHourString'] = str(int(eachday['opening']/60)).zfill(2)+':'+str(eachday['opening']%60).zfill(2)
		day_json['closingHourString'] = str(int(eachday['closing']/60)).zfill(2)+':'+str(eachday['closing']%60).zfill(2)
		#print('to be uploaded into db:',day_json)
		operatingHours.append(day_json)
	return operatingHours

class JSONEncoder(json.JSONEncoder):
	def default(self, o):
		if isinstance(o, ObjectId):
			return str(o)
		if isinstance(o, datetime.datetime):
			return str(o)
		return json.JSONEncoder.default(self, o)

def is_more_than_one_for_day(operatingHours_array):
	counter = 0
	myarr=[0,0,0,0,0,0,0]
	for eachday in operatingHours_array:
		myarr[eachday['dayOfTheWeek']-1] += 1
	print(myarr)	
	return max(myarr)
		
	
client = pymongo.MongoClient('ds051770-a0.mongolab.com',51770)
db = client.heroku_app31071968
db.authenticate('Bora_write', 'MmPasa_189m')#readonly user
db_venues = db.venues

start = time.time()

all_db_venues = list(db_venues.find({ '$query': {"creationDate":{'$lt':datetime.datetime(2014, 11, 20)}}},{'operatingHours':1,'name':1,'_id':1}))
#print ("all db venues:",all_db_venues)
#total_n_of_venue = db_venues.count()


def checkfun(all_db_venues):
	for dbvenue in all_db_venues:
		dbname = dbvenue['name']
		operatingHours_array = dbvenue['operatingHours']
		day_1_counter = is_more_than_one_for_day(operatingHours_array)
		if day_1_counter!=1:
			print(dbvenue['_id'],dbname,day_1_counter)

checkfun(all_db_venues)
end = time.time()

print (end - start)
