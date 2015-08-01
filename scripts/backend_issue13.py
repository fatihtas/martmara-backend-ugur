import json
import pymongo
import sys
import datetime, time
import bson
import os
import openpyxl
from bson.objectid import ObjectId
from bson import Binary, Code
from bson import json_util
from difflib import SequenceMatcher
#MmPasa_189m

def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

def get_venueobjarray():
	wb = openpyxl.reader.excel.load_workbook('Viveka Transferred Data.xlsx',data_only = True,guess_types=True)
	#print (wb.get_sheet_names())
	sheet_ranges = wb['Data Entry Sheet']
	allvenues = []
	for x in range(24,2122):#for each venue
		rowtuple = {}
		name = sheet_ranges['H'+str(x)].value#[,[]]
		while name.endswith(' '):
			name = name[:-1]
		rowtuple['name'] = name
		rowtuple['workinghours'] = []
		for day in range(7):#for each day
			opening = sheet_ranges[openpyxl.utils.get_column_letter(day*2+16)+str(x)].value
			closing = sheet_ranges[openpyxl.utils.get_column_letter(day*2+17)+str(x)].value
			
			if (not isinstance(opening,datetime.datetime) and not isinstance(opening,datetime.time)):
				print("bugun calismiyoruz.",opening,closing)
				continue
			if isinstance(opening,datetime.datetime):
				opening = opening.time()
			if isinstance(closing,datetime.datetime):
				closing = closing.time()	
			opening = opening.hour*60+opening.minute
			closing = closing.hour*60+closing.minute
			
			rowtuple['workinghours'].append({'day':day+1,'opening':opening,'closing':closing})
		allvenues.append(rowtuple)
	return allvenues
	
def update(db_venues,day_json,_id,dbname,excelvenuename):
	print ("updating: ",_id,'\t',dbname,'\t',excelvenuename)
	print ("with json:",day_json)
	db_venues.update({
		'_id' : _id
		},{
		'$set':{
		'operatingHours':day_json
		}
		}
		,multi=False, upsert=False
		)
		
	return
		
	
def get_day_json(correctvenue):
	operatingHours = []
	for eachday in correctvenue['workinghours']:
		day_json={}
		day_json['dayOfTheWeek'] = eachday['day']
		day_json['openingHour'] = eachday['opening']
		day_json['closingHour'] = eachday['closing']
		day_json['openingHourString'] = str(int(eachday['opening']/60)).zfill(2)+':'+str(eachday['opening']%60).zfill(2)
		day_json['closingHourString'] = str(int(eachday['closing']/60)).zfill(2)+':'+str(eachday['closing']%60).zfill(2)
		#print('to be uploaded into db:',day_json)
		operatingHours.append(day_json)
	return operatingHours

#buradaki araligi degistirerek venue set'ten select yapabiliriz
#ornek: _skip=100 _limit=200 olursa [100,300] venue'larinda islem yapilir.
#_skip = int(input("start:"))#start index
#_limit = int(input("limit:"))#number of venues to write to file

class JSONEncoder(json.JSONEncoder):
	def default(self, o):
		if isinstance(o, ObjectId):
			return str(o)
		if isinstance(o, datetime.datetime):
			return str(o)
		return json.JSONEncoder.default(self, o)



client = pymongo.MongoClient('ds051770-a0.mongolab.com',51770)
db = client.heroku_app31071968
db.authenticate('Bora_write', 'MmPasa_189m')#readonly user
db_venues = db.venues

dbdebulundu_count=0

start = time.time()

all_db_venues = list(db_venues.find({ '$query': {"creationDate":{'$lt':datetime.datetime(2014, 11, 20)}}},{'name':1,'_id':1}))
#print ("all db venues:",all_db_venues)
#total_n_of_venue = db_venues.count()
all_excel_venues = get_venueobjarray()
try:
    os.remove('not_added_venues.txt')
except OSError:
    pass
file_not_added_venue_names = open('not_added_venues.txt', 'a')

for correctvenue in all_excel_venues:
	#print ("Json:",eachvenue)
	excelvenuename = correctvenue['name']
	#print ("venue name in excel:",excelvenuename)
	for dbvenue in all_db_venues:
		dbname = dbvenue['name']
		similarity = similar(excelvenuename,dbname)
		if similarity>0.9 and similarity<1.0:
			print ('\n',dbname,'\t',excelvenuename)
			decision = input("what to do with this two(y,n): ")
			if(decision=='y'):
				day_json = get_day_json(correctvenue)
				update(db_venues,day_json,dbvenue['_id'],dbname,excelvenuename)
				dbvenue['updated']=1
				break
			#print(similarity, dbname,excelvenuename)
		elif similarity==1.0:
			day_json = get_day_json(correctvenue)
			update(db_venues,day_json,dbvenue['_id'],dbname,excelvenuename)
			dbvenue['updated']=1
			break
	"""thisvenuefound=0
	for venue_db in all_db_venues:
		#print (venue_db)
		_id = venue_db['_id']
		print('_id in db:',_id)
		dbdebulundu_count += 1
		thisvenuefound+=1
		if(thisvenuefound>1):
			print("!!!!!this venue is found more than one!!!!!")
		
		operatingHours=[]
		for eachday in correctvenue['workinghours']:
			#print (eachday)
			day_json['dayOfTheWeek'] = eachday['day']
			day_json['openingHour'] = eachday['opening']
			day_json['closingHour'] = eachday['closing']
			day_json['openingHourString'] = str(int(eachday['opening']/60)).zfill(2)+':'+str(eachday['opening']%60).zfill(2)
			day_json['closingHourString'] = str(int(eachday['closing']/60)).zfill(2)+':'+str(eachday['closing']%60).zfill(2)
			#print('to be uploaded into db:',day_json)
			operatingHours.append(day_json)
		print('operatingHours json: ',operatingHours)
		db_venues.update({
		'_id' : _id
		},{
		'$set':{
		'operatingHours':operatingHours
		}
		}
		,multi=False, upsert=False
		)
		print('')
	print('')
	#print ("bulundu")"""
	
		#operatingHours.append
		#print("Db Object:",dbobject)
""""	#for eachWorkingInterval in eachvenue:
	#	day_dict = {}
	#	day_dict['dayOfTheWeek'] = eachWorkingInterval['day']
		#print (type(eachvenue[1][0]))
		if ((type(eachvenue[1][eachday*2]) is not datetime.time) and (type(eachvenue[1][eachday*2]) is not datetime.datetime) and (type(eachvenue[1][(eachday*2)+1]) is not datetime.time) and (type(eachvenue[1][(eachday*2)+1]) is not datetime.datetime)):
			print (eachvenue, eachday+1)
			continue
		day_dict = {}
		day_dict['dayOfTheWeek']=eachday+1
		day_dict['openingHourString']=str(eachvenue[1][eachday*2])
		day_dict['closingHourString']=eachday+1
		day_dict['openingHour']=eachday+1
		day_dict['closingHour']=eachday+1
		#print (day_dict['openingHourString'])
		#operatingHours.append()
	
	#print(type(x[1][1]))"""

"""for venue in db_venues.find({},{'_id':1,'name':1,'operatingHours':1}).skip(_skip).limit(_limit):

	venue_name = venue['name']
	

	
	
	#below finds the products of this venue
	products_of_this_venue = []
	product_id_array = []
	for productofvenue in db_venueproducts.find({'venue':venue_id},{'product':1,'rating':1,'price':1,'currency':1}):
		products_of_this_venue.append(productofvenue)#(productofvenue['product'],productofvenue['rating'],productofvenue['price'],productofvenue['currency']))
		product_id_array.append(productofvenue['product'])#for teh products query.
	products = list(db_products.find({'_id':{'$in' : product_id_array}},{'nameEn':1,'nameTr':1,'_id':1}))
	#burada nested loop yapmak zorundayim, python'da dictionary'leri join edince de ayni seyi yapiyomus. n=number_of_total_venues ise bizim kod n^3'te calisiyor.
	for product in products_of_this_venue:
		for productofvenue in products:
			if product['product']==productofvenue['_id']:
				if 'nameEn' in productofvenue:
					product['nameEn']=productofvenue['nameEn']
				if 'nameTr' in productofvenue:
					product['nameTr']=productofvenue['nameTr']
		product.pop('_id',None)
		product.pop('product',None)
		
	#print (products_of_this_venue)
	
	venue['products'] = products_of_this_venue
	myc+=1
	
	if (commatag==0):
		f.write(bytes(str(JSONEncoder().encode(venue)),'utf-8'))
		commatag=1
	else:
		f.write(bytes(',\n'+str(JSONEncoder().encode(venue)),'utf-8'))
	#print ("venue:",myc,"\ttotalvenue:", total_n_of_venue,"\tproductintisvenue:",len(product_id_array))

"""

end = time.time()

print (end - start)
